"""
Parser inteligente para formularios R-60
Identifica automáticamente el tipo de formulario y extrae datos estructurados
"""

from pathlib import Path
from typing import Dict, List, Optional, Any
from datetime import datetime

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

import config
from src.utils.logger import logger, log_execution_time
from src.utils.exceptions import (
    ArchivoExcelInvalidoError,
    TipoFormularioNoReconocidoError,
    CampoObligatorioError,
    ItemsVaciosError
)


class ExcelParser:
    """Parser para formularios R-60 en formato Excel"""
    
    def __init__(self, file_path: Path):
        """
        Inicializa el parser
        
        Args:
            file_path: Ruta al archivo Excel
        """
        self.file_path = file_path
        self.workbook: Optional[Workbook] = None
        self.sheet: Optional[Worksheet] = None
        self.form_type: Optional[str] = None
        
    @log_execution_time(logger)
    def parse(self) -> Dict[str, Any]:
        """
        Parsea el formulario completo
        
        Returns:
            Diccionario con todos los datos estructurados del formulario
            
        Raises:
            ArchivoExcelInvalidoError: Si el archivo está corrupto
            TipoFormularioNoReconocidoError: Si no se puede identificar el tipo
            CampoObligatorioError: Si falta un campo obligatorio
            ItemsVaciosError: Si no hay ítems para procesar
        """
        logger.info(f"Parseando formulario: {self.file_path.name}")
        
        # 1. Abrir el workbook
        self._load_workbook()
        
        # 2. Identificar tipo de formulario
        self.form_type = self._identify_form_type()
        logger.info(f"Tipo de formulario identificado: {self.form_type}")
        
        # 3. Extraer datos del encabezado
        header_data = self._extract_header_data()
        
        # 4. Extraer ítems
        items_data = self._extract_items_data()
        
        # 5. Validar que haya ítems
        if not items_data:
            raise ItemsVaciosError(header_data.get('numero_formulario', 'Desconocido'))
        
        # 6. Construir resultado
        result = {
            **header_data,
            'tipo_formulario': self.form_type,
            'items': items_data,
            'archivo_original': self.file_path.name
        }
        
        logger.info(f"✅ Formulario parseado: {len(items_data)} ítem(s) extraído(s)")
        
        # Cerrar workbook
        self.workbook.close()
        
        return result
    
    def _load_workbook(self) -> None:
        """
        Carga el workbook de Excel
        
        Raises:
            ArchivoExcelInvalidoError: Si el archivo no se puede abrir
        """
        try:
            self.workbook = openpyxl.load_workbook(
                self.file_path, 
                data_only=True  # Leer valores calculados en lugar de fórmulas
            )
            # Usar la primera hoja activa
            self.sheet = self.workbook.active
            logger.debug(f"Workbook cargado: {self.sheet.title}")
        
        except Exception as e:
            raise ArchivoExcelInvalidoError(str(self.file_path), str(e))
    
    def _identify_form_type(self) -> str:
        """
        Identifica automáticamente el tipo de formulario
        
        Busca palabras clave en:
        1. Nombre de la hoja
        2. Contenido de celdas clave
        
        Returns:
            Tipo de formulario: COMPRAS, SERVICIOS o COSTOS
            
        Raises:
            TipoFormularioNoReconocidoError: Si no se puede identificar
        """
        # Buscar en el nombre de la hoja
        sheet_name = self.sheet.title.lower()
        
        for form_type, keywords in config.FORM_TYPE_KEYWORDS.items():
            for keyword in keywords:
                if keyword.lower() in sheet_name:
                    logger.debug(f"Tipo identificado por nombre de hoja: {form_type}")
                    return form_type
        
        # Buscar en contenido de celdas (primeras 10 filas)
        content = ""
        for row in self.sheet.iter_rows(min_row=1, max_row=10, values_only=True):
            content += " ".join(str(cell).lower() for cell in row if cell)
        
        for form_type, keywords in config.FORM_TYPE_KEYWORDS.items():
            for keyword in keywords:
                if keyword.lower() in content:
                    logger.debug(f"Tipo identificado por contenido: {form_type}")
                    return form_type
        
        # Si no se identifica, lanzar excepción
        raise TipoFormularioNoReconocidoError(str(self.file_path))
    
    def _extract_header_data(self) -> Dict[str, Any]:
        """
        Extrae los datos del encabezado del formulario
        
        Returns:
            Diccionario con los datos del encabezado
            
        Raises:
            CampoObligatorioError: Si falta un campo obligatorio
        """
        mapping = config.FORM_TYPE_MAPPINGS[self.form_type]['header']
        header_data = {}
        
        for field_name, cell_ref in mapping.items():
            value = self.sheet[cell_ref].value
            
            # Limpiar y convertir valor
            if value is not None:
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d')
                else:
                    value = str(value).strip()
            else:
                value = ""
            
            header_data[field_name] = value
            logger.debug(f"{field_name} ({cell_ref}): {value}")
        
        # Validar campos obligatorios
        campos_obligatorios = ['numero_formulario', 'solicitante']
        
        for campo in campos_obligatorios:
            if not header_data.get(campo):
                celda = mapping.get(campo, 'desconocida')
                raise CampoObligatorioError(campo, celda)
        
        logger.debug("Datos de encabezado extraídos correctamente")
        return header_data
    
    def _extract_items_data(self) -> List[Dict[str, Any]]:
        """
        Extrae los datos de la tabla de ítems
        
        Returns:
            Lista de diccionarios con los datos de cada ítem
        """
        items_start_row = config.FORM_TYPE_MAPPINGS[self.form_type]['items_start_row']
        columns_mapping = config.FORM_TYPE_MAPPINGS[self.form_type]['items_columns']
        
        items = []
        current_row = items_start_row
        
        logger.debug(f"Extrayendo ítems desde fila {items_start_row}")
        
        while True:
            # Leer el número de ítem de la primera columna
            item_number_col = columns_mapping.get('numero_item', 'A')
            item_number = self.sheet[f"{item_number_col}{current_row}"].value
            
            # Si el número de ítem está vacío, terminamos
            if not item_number or str(item_number).strip() == "":
                logger.debug(f"Fin de ítems en fila {current_row}")
                break
            
            # Extraer datos del ítem según el tipo de formulario
            item_data = self._extract_item_row(current_row, columns_mapping)
            
            if item_data:
                items.append(item_data)
                logger.debug(f"Ítem {len(items)} extraído: {item_data.get('numero_item')}")
            
            current_row += 1
            
            # Protección contra loops infinitos (máximo 1000 filas)
            if current_row > items_start_row + 1000:
                logger.warning("Se alcanzó el límite de 1000 filas")
                break
        
        logger.debug(f"Total de ítems extraídos: {len(items)}")
        return items
    
    def _extract_item_row(self, row: int, columns_mapping: Dict[str, str]) -> Dict[str, Any]:
        """
        Extrae los datos de una fila de ítem
        
        Args:
            row: Número de fila
            columns_mapping: Mapeo de columnas para este tipo de formulario
            
        Returns:
            Diccionario con los datos del ítem
        """
        item = {}
        
        # Extraer todos los campos según el mapeo del config
        for field_name, column in columns_mapping.items():
            cell_ref = f"{column}{row}"
            item[field_name] = self._get_cell_value(cell_ref)
        
        return item
    
    def _get_cell_value(self, cell_ref: str) -> Any:
        """
        Obtiene el valor de una celda y lo formatea apropiadamente
        
        Args:
            cell_ref: Referencia de la celda (ej: 'A1')
            
        Returns:
            Valor formateado de la celda
        """
        value = self.sheet[cell_ref].value
        
        if value is None:
            return ""
        
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')
        
        if isinstance(value, (int, float)):
            return value
        
        return str(value).strip()


def parse_r60_form(file_path: Path) -> Dict[str, Any]:
    """
    Función de conveniencia para parsear un formulario R-60
    
    Args:
        file_path: Ruta al archivo Excel
        
    Returns:
        Datos estructurados del formulario
    """
    parser = ExcelParser(file_path)
    return parser.parse()


if __name__ == "__main__":
    # Test del parser
    import sys
    
    if len(sys.argv) < 2:
        print("Uso: python excel_parser.py <ruta_al_excel>")
        sys.exit(1)
    
    test_file = Path(sys.argv[1])
    
    if not test_file.exists():
        print(f"❌ Archivo no encontrado: {test_file}")
        sys.exit(1)
    
    try:
        result = parse_r60_form(test_file)
        
        print("\n✅ Formulario parseado exitosamente!")
        print(f"\nTipo: {result['tipo_formulario']}")
        print(f"Número: {result['numero_formulario']}")
        print(f"Solicitante: {result['solicitante']}")
        print(f"Ítems: {len(result['items'])}")
        print("\nPrimeros 3 ítems:")
        for i, item in enumerate(result['items'][:3], 1):
            print(f"  {i}. {item}")
    
    except Exception as e:
        print(f"❌ Error al parsear: {e}")
        import traceback
        traceback.print_exc()


