#!/usr/bin/env python3
"""
Script para crear formularios R-60 de prueba
Útil para testing sin necesidad de formularios reales
"""

import argparse
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def create_compras_form(numero: str, solicitante: str, output_path: Path):
    """Crea un formulario de COMPRAS de prueba"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulario de Compras"
    
    # Título
    ws['A1'] = "FORMULARIO R-60 - COMPRAS"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Encabezado
    ws['C2'] = "Nº de Formulario:"
    ws['D2'] = numero
    ws['C3'] = "Fecha:"
    ws['D3'] = datetime.now().strftime('%Y-%m-%d')
    ws['C4'] = "Solicitante:"
    ws['D4'] = solicitante
    ws['C5'] = "Área:"
    ws['D5'] = "Departamento de Pruebas"
    ws['C6'] = "Observaciones:"
    ws['D6'] = "Formulario generado automáticamente para pruebas"
    
    # Encabezado de tabla
    headers = ['Nº Item', 'Descripción', 'Cantidad', 'Unidad', 'Precio Unit.', 'Total']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Ítems de ejemplo
    items = [
        (1, "Laptop HP ProBook", 2, "unid", 1200.00, 2400.00),
        (2, "Mouse Logitech", 5, "unid", 25.00, 125.00),
        (3, "Teclado Mecánico", 3, "unid", 80.00, 240.00),
        (4, "Monitor 24 pulgadas", 2, "unid", 350.00, 700.00),
    ]
    
    for i, item in enumerate(items, start=10):
        ws.cell(row=i, column=1).value = item[0]
        ws.cell(row=i, column=2).value = item[1]
        ws.cell(row=i, column=3).value = item[2]
        ws.cell(row=i, column=4).value = item[3]
        ws.cell(row=i, column=5).value = item[4]
        ws.cell(row=i, column=6).value = item[5]
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    wb.save(output_path)
    print(f"✅ Formulario de COMPRAS creado: {output_path}")


def create_servicios_form(numero: str, solicitante: str, output_path: Path):
    """Crea un formulario de SERVICIOS de prueba"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulario de Servicios"
    
    # Título
    ws['A1'] = "FORMULARIO R-60 - SERVICIOS"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Encabezado
    ws['C2'] = "Nº de Formulario:"
    ws['D2'] = numero
    ws['C3'] = "Fecha:"
    ws['D3'] = datetime.now().strftime('%Y-%m-%d')
    ws['C4'] = "Solicitante:"
    ws['D4'] = solicitante
    ws['C5'] = "Área:"
    ws['D5'] = "Departamento de Servicios"
    ws['C6'] = "Observaciones:"
    ws['D6'] = "Servicios contratados - Formulario de prueba"
    
    # Encabezado de tabla
    headers = ['Nº Item', 'Servicio', 'Proveedor', 'Monto', 'Fecha Servicio']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Ítems de ejemplo
    items = [
        (1, "Mantenimiento de servidores", "Tech Solutions SA", 1500.00, "2024-01-15"),
        (2, "Consultoría IT", "Consultores Pro", 2500.00, "2024-01-20"),
        (3, "Soporte técnico mensual", "Support Inc", 800.00, "2024-01-01"),
    ]
    
    for i, item in enumerate(items, start=10):
        ws.cell(row=i, column=1).value = item[0]
        ws.cell(row=i, column=2).value = item[1]
        ws.cell(row=i, column=3).value = item[2]
        ws.cell(row=i, column=4).value = item[3]
        ws.cell(row=i, column=5).value = item[4]
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    
    wb.save(output_path)
    print(f"✅ Formulario de SERVICIOS creado: {output_path}")


def create_costos_form(numero: str, solicitante: str, output_path: Path):
    """Crea un formulario de COSTOS de prueba"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulario de Costos"
    
    # Título
    ws['A1'] = "FORMULARIO R-60 - COSTOS"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Encabezado
    ws['C2'] = "Nº de Formulario:"
    ws['D2'] = numero
    ws['C3'] = "Fecha:"
    ws['D3'] = datetime.now().strftime('%Y-%m-%d')
    ws['C4'] = "Solicitante:"
    ws['D4'] = solicitante
    ws['C5'] = "Área:"
    ws['D5'] = "Departamento de Finanzas"
    ws['C6'] = "Observaciones:"
    ws['D6'] = "Registro de gastos operativos - Prueba"
    
    # Encabezado de tabla
    headers = ['Nº Item', 'Concepto', 'Categoría', 'Monto', 'Fecha']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Ítems de ejemplo
    items = [
        (1, "Papelería y útiles", "Materiales", 350.00, "2024-01-10"),
        (2, "Servicios públicos", "Operativo", 1200.00, "2024-01-05"),
        (3, "Combustible vehículos", "Transporte", 600.00, "2024-01-12"),
        (4, "Capacitación personal", "RRHH", 2000.00, "2024-01-18"),
    ]
    
    for i, item in enumerate(items, start=10):
        ws.cell(row=i, column=1).value = item[0]
        ws.cell(row=i, column=2).value = item[1]
        ws.cell(row=i, column=3).value = item[2]
        ws.cell(row=i, column=4).value = item[3]
        ws.cell(row=i, column=5).value = item[4]
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    wb.save(output_path)
    print(f"✅ Formulario de COSTOS creado: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description='Crear formularios R-60 de prueba',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Ejemplos de uso:
  python create_test_form.py compras -n R60-001 -s "Juan Pérez"
  python create_test_form.py servicios -n R60-002 -s "María García"
  python create_test_form.py costos -n R60-003 -s "Carlos López"
  python create_test_form.py all  # Crea los 3 tipos
        '''
    )
    
    parser.add_argument(
        'tipo',
        choices=['compras', 'servicios', 'costos', 'all'],
        help='Tipo de formulario a crear'
    )
    
    parser.add_argument(
        '-n', '--numero',
        default=f"TEST-{datetime.now().strftime('%Y%m%d-%H%M')}",
        help='Número de formulario (default: TEST-YYYYMMDD-HHMM)'
    )
    
    parser.add_argument(
        '-s', '--solicitante',
        default='Usuario de Prueba',
        help='Nombre del solicitante (default: Usuario de Prueba)'
    )
    
    parser.add_argument(
        '-o', '--output',
        type=Path,
        default=Path('temp'),
        help='Carpeta de salida (default: temp/)'
    )
    
    args = parser.parse_args()
    
    # Crear carpeta de salida si no existe
    args.output.mkdir(parents=True, exist_ok=True)
    
    print(f"\n🏭 Generando formulario(s) R-60 de prueba...\n")
    
    if args.tipo == 'compras' or args.tipo == 'all':
        numero = f"{args.numero}-COMPRAS" if args.tipo == 'all' else args.numero
        output_file = args.output / f"R60_Compras_{numero.replace('/', '-')}.xlsx"
        create_compras_form(numero, args.solicitante, output_file)
    
    if args.tipo == 'servicios' or args.tipo == 'all':
        numero = f"{args.numero}-SERVICIOS" if args.tipo == 'all' else args.numero
        output_file = args.output / f"R60_Servicios_{numero.replace('/', '-')}.xlsx"
        create_servicios_form(numero, args.solicitante, output_file)
    
    if args.tipo == 'costos' or args.tipo == 'all':
        numero = f"{args.numero}-COSTOS" if args.tipo == 'all' else args.numero
        output_file = args.output / f"R60_Costos_{numero.replace('/', '-')}.xlsx"
        create_costos_form(numero, args.solicitante, output_file)
    
    print(f"\n✅ Formulario(s) creado(s) en: {args.output}/")
    print(f"\nPuedes probarlos con:")
    print(f"  python run.py test-parser {args.output}/R60_*.xlsx")


if __name__ == '__main__':
    main()


